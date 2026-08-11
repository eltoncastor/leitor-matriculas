"""
xlsx_exporter.py

Gera a planilha XLSX final (abas Liberações, Revisão, Resumo) a partir de
uma lista de dicionários já prontos (montados em ui.py a partir dos
Registro do registro_parser + validacao.classificar_registro). Não lê nem
inventa dado nenhum aqui — só formata, ordena e escreve.

ESTRUTURA DA SAÍDA (requisito funcional definitivo): as 7 primeiras
colunas seguem sempre esta ordem fixa — Data, Hora, Matrícula, Nome,
Setor, Motivo, Responsável pela autorização. As demais colunas (Cargo,
Página, Status, confianças, Observação, texto OCR original) são
informação técnica/auditoria, mantida depois das 7 principais.

ORDEM (requisito funcional definitivo): a ordem FÍSICA das liberações na
folha é preservada. Este módulo NÃO reordena nada — grava exatamente na
ordem em que os registros chegam, que é a ordem de leitura da página (o
parser espacial devolve as linhas de cima para baixo) e das páginas entre
si. Não ordena por data/hora, nome, matrícula, gestor nem motivo: a
sequência física do papel é a referência da saída, e é o que permite
conferir a planilha contra a folha linha a linha.
"""

import os
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUNAS = [
    ("data", "Data"),
    ("hora", "Hora"),
    ("matricula", "Matrícula"),
    ("nome", "Nome"),
    ("setor", "Setor"),
    ("motivo", "Motivo"),
    ("gestor", "Responsável pela Autorização"),
    ("cargo", "Cargo"),
    ("pagina_origem", "Página"),
    ("status", "Status"),
    ("confianca_matricula", "Confiança Matrícula"),
    ("confianca_gestor", "Confiança Responsável"),
    ("confianca_motivo", "Confiança Motivo"),
    ("observacao", "Observação"),
    ("texto_ocr_original", "Texto OCR (matrícula)"),
]

_IDX_MATRICULA = [i for i, (chave, _) in enumerate(COLUNAS) if chave == "matricula"][0]


def _fmt_confianca(v):
    return f"{v * 100:.1f}%" if isinstance(v, (int, float)) else (v or "")


def _cabecalho(ws):
    ws.append([rotulo for _, rotulo in COLUNAS])
    for cel in ws[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="4472C4")
        cel.alignment = Alignment(horizontal="center")
    for i, (_, rotulo) in enumerate(COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(rotulo) + 4)


def _escrever_linhas(ws, registros: List[Dict]):
    for r in registros:
        linha = []
        for chave, _ in COLUNAS:
            v = r.get(chave, "")
            if chave.startswith("confianca"):
                v = _fmt_confianca(v)
            linha.append(v if v is not None else "")
        ws.append(linha)

    col_letra = get_column_letter(_IDX_MATRICULA + 1)
    for linha_num in range(2, ws.max_row + 1):
        cel = ws[f"{col_letra}{linha_num}"]
        cel.number_format = "@"  # texto -- nunca deixar o Excel converter a matrícula em número
        if cel.value is not None:
            cel.value = str(cel.value)


def _aplicar_tabela(ws, nome_tabela: str):
    if ws.max_row < 2:
        return
    try:
        ultima_coluna = get_column_letter(len(COLUNAS))
        tabela = Table(displayName=nome_tabela, ref=f"A1:{ultima_coluna}{ws.max_row}")
        tabela.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tabela)
    except Exception:
        pass  # formatação é acessório -- nunca deve derrubar a exportação
    ws.freeze_panes = "A2"


def _aba_resumo(wb, registros, paginas_processadas, paginas_com_erro, paginas_com_contagem_divergente=0):
    ws = wb.create_sheet("Resumo")
    ws.append(["Resumo"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    confirmados = sum(1 for r in registros if r.get("status") == "CONFIRMADO")
    revisao = sum(1 for r in registros if r.get("status") == "REVISAO")
    erro = sum(1 for r in registros if r.get("status") == "ERRO")

    for rotulo, valor in [
        ("Total de liberações", len(registros)),
        ("Confirmadas", confirmados),
        ("Em revisão", revisao),
        ("Erros", erro),
        ("Páginas processadas", paginas_processadas),
        ("Páginas com erro", paginas_com_erro),
        # PROBLEMA 2: 8 posições esperadas por folha -- só informativo,
        # nunca força nem descarta registro (ver registro_parser.
        # verificar_contagem_posicoes).
        ("Páginas com contagem de posições divergente do esperado", paginas_com_contagem_divergente),
    ]:
        ws.append([rotulo, valor])

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12

    for titulo, chave in [("Liberações por gestor", "gestor"), ("Liberações por motivo", "motivo")]:
        ws.append([])
        ws.append([titulo])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        contagem: Dict[str, int] = {}
        for r in registros:
            k = r.get(chave) or "(não identificado)"
            contagem[k] = contagem.get(k, 0) + 1
        for k, qtd in sorted(contagem.items(), key=lambda item: -item[1]):
            ws.append([k, qtd])


def export_to_xlsx(
    registros: List[Dict],
    caminho: str,
    paginas_processadas: int = 0,
    paginas_com_erro: int = 0,
    paginas_com_contagem_divergente: int = 0,
) -> None:
    """
    registros: lista de dicts com as chaves de COLUNAS (mais 'status' com
    valor "CONFIRMADO"/"REVISAO"/"ERRO"). Cria/sobrescreve o arquivo em
    `caminho` com 3 abas: Liberações (tudo, na ordem física da folha),
    Revisão (REVISAO + ERRO, na mesma ordem física), Resumo (contagens).

    A ordem de `registros` é preservada exatamente como recebida — ver a
    seção ORDEM no topo do módulo.
    """
    if not registros:
        raise ValueError("Nenhum registro para exportar.")

    diretorio = os.path.dirname(os.path.abspath(caminho))
    if diretorio and not os.path.isdir(diretorio):
        raise FileNotFoundError(f"Pasta de destino não encontrada: {diretorio}")

    wb = Workbook()
    ws_lib = wb.active
    ws_lib.title = "Liberações"
    _cabecalho(ws_lib)
    _escrever_linhas(ws_lib, registros)
    _aplicar_tabela(ws_lib, "TabelaLiberacoes")

    ws_rev = wb.create_sheet("Revisão")
    _cabecalho(ws_rev)
    problematicos = [r for r in registros if r.get("status") in ("REVISAO", "ERRO")]
    _escrever_linhas(ws_rev, problematicos)
    _aplicar_tabela(ws_rev, "TabelaRevisao")

    _aba_resumo(wb, registros, paginas_processadas, paginas_com_erro, paginas_com_contagem_divergente)

    try:
        wb.save(caminho)
    except PermissionError as exc:
        raise PermissionError(
            f"Não foi possível salvar '{caminho}' — feche o arquivo se ele já estiver aberto no Excel/LibreOffice."
        ) from exc
