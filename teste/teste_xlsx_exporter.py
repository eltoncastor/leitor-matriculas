import os
import sys
import shutil
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402
from xlsx_exporter import export_to_xlsx  # noqa: E402


def _registro(status, matricula="28972", pagina=1, **extra):
    r = {
        "data": "23/04/2026", "hora": "11:05", "matricula": matricula, "nome": "Fulano",
        "cargo": "Analista", "setor": "TI", "gestor": "Gestor X", "motivo": "RH",
        "pagina_origem": pagina, "status": status,
        "confianca_matricula": 0.95, "confianca_gestor": 0.9, "confianca_motivo": 0.9,
        "observacao": "", "texto_ocr_original": matricula,
    }
    r.update(extra)
    return r


def teste_export_basico():
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida.xlsx")
        registros = [
            _registro("CONFIRMADO", matricula="00312"),
            _registro("REVISAO", matricula="99999", observacao="matrícula não encontrada"),
            _registro("ERRO", matricula="", observacao="falha ao processar página"),
        ]
        export_to_xlsx(registros, caminho, paginas_processadas=5, paginas_com_erro=1)

        wb = openpyxl.load_workbook(caminho)
        assert wb.sheetnames == ["Liberações", "Revisão", "Resumo"], wb.sheetnames

        ws_lib = wb["Liberações"]
        assert ws_lib.max_row == 4  # cabeçalho + 3 registros

        ws_rev = wb["Revisão"]
        assert ws_rev.max_row == 3  # cabeçalho + REVISAO + ERRO (não o CONFIRMADO)

        print("OK: 3 abas criadas, Liberações com todos, Revisão só com REVISAO/ERRO")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teste_matricula_como_texto_preserva_zero():
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida2.xlsx")
        export_to_xlsx([_registro("CONFIRMADO", matricula="00312")], caminho)

        wb = openpyxl.load_workbook(caminho)
        ws = wb["Liberações"]
        idx_matricula_col = None
        for cel in ws[1]:
            if cel.value == "Matrícula":
                idx_matricula_col = cel.column
        celula = ws.cell(row=2, column=idx_matricula_col)
        assert celula.value == "00312", celula.value
        assert celula.number_format == "@"
        print("OK: matrícula '00312' preservada como texto no XLSX, formato '@'")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teste_resumo_contagens():
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida3.xlsx")
        registros = [_registro("CONFIRMADO"), _registro("CONFIRMADO"), _registro("REVISAO")]
        export_to_xlsx(registros, caminho, paginas_processadas=2, paginas_com_erro=0)

        wb = openpyxl.load_workbook(caminho)
        ws = wb["Resumo"]
        valores = {row[0].value: row[1].value for row in ws.iter_rows(min_row=3, max_row=8) if row[0].value}
        assert valores["Total de liberações"] == 3
        assert valores["Confirmadas"] == 2
        assert valores["Em revisão"] == 1
        assert valores["Páginas processadas"] == 2
        print("OK: aba Resumo com contagens corretas")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teste_sem_registros():
    try:
        export_to_xlsx([], "/tmp/nao_importa.xlsx")
        raise AssertionError("deveria levantar ValueError")
    except ValueError:
        print("OK: lista vazia -> ValueError")


def teste_pasta_inexistente():
    try:
        export_to_xlsx([_registro("CONFIRMADO")], "/caminho/que/nao/existe/saida.xlsx")
        raise AssertionError("deveria levantar FileNotFoundError")
    except FileNotFoundError:
        print("OK: pasta de destino inexistente -> FileNotFoundError")


def teste_ordem_das_colunas_principais():
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida_colunas.xlsx")
        export_to_xlsx([_registro("CONFIRMADO")], caminho)
        wb = openpyxl.load_workbook(caminho)
        ws = wb["Liberações"]
        cabecalhos = [cel.value for cel in ws[1]]
        # Requisito funcional definitivo: as 7 primeiras colunas devem ser
        # exatamente esta ordem -- Data, Hora, Matrícula, Nome, Setor,
        # Motivo, Responsável.
        assert cabecalhos[:7] == [
            "Data", "Hora", "Matrícula", "Nome", "Setor", "Motivo",
            "Responsável pela Autorização",
        ], cabecalhos[:7]
        print("OK: 7 primeiras colunas na ordem exigida (Data/Hora/Matrícula/Nome/Setor/Motivo/Responsável)")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teste_ordem_fisica_da_folha_e_preservada():
    """
    REQUISITO FUNCIONAL DEFINITIVO: a ordem física das liberações na folha
    é preservada. O exportador NÃO reordena cronologicamente -- data/hora
    fora de ordem cronológica saem exatamente na ordem em que chegaram.
    """
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida_ordem_fisica.xlsx")
        # Datas/horas deliberadamente fora de ordem cronológica: uma
        # ordenação por data+hora produziria [pagina2, pagina4, pagina3,
        # pagina1]. A ordem de entrada tem de ser mantida intacta.
        registros = [
            _registro("CONFIRMADO", matricula="pagina1", data="15/08/2026", hora="16:10"),
            _registro("CONFIRMADO", matricula="pagina2", data="14/08/2026", hora="08:15"),
            _registro("CONFIRMADO", matricula="pagina3", data="15/08/2026", hora="07:42"),
            _registro("CONFIRMADO", matricula="pagina4", data="14/08/2026", hora="14:20"),
        ]
        export_to_xlsx(registros, caminho)

        wb = openpyxl.load_workbook(caminho)
        ws = wb["Liberações"]
        # Coluna 3 = Matrícula (Data, Hora, Matrícula, ...)
        matriculas_na_ordem = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        assert matriculas_na_ordem == ["pagina1", "pagina2", "pagina3", "pagina4"], matriculas_na_ordem
        print("OK: ordem física da folha preservada (sem reordenação cronológica)")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teste_registro_sem_hora_mantem_sua_posicao():
    """
    Com a HORA opcional, um registro sem hora é comum. Ele não pode ser
    empurrado para o fim da planilha: mantém a posição física que tinha na
    folha.
    """
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida_sem_hora.xlsx")
        registros = [
            _registro("CONFIRMADO", matricula="sem-hora", data="14/08/2026", hora=""),
            _registro("CONFIRMADO", matricula="com-hora", data="14/08/2026", hora="08:15"),
            _registro("REVISAO", matricula="sem-data-nem-hora", data="", hora=""),
        ]
        export_to_xlsx(registros, caminho)

        wb = openpyxl.load_workbook(caminho)
        ws = wb["Liberações"]
        matriculas_na_ordem = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        # Nada é reordenado nem descartado: a ordem de entrada é a saída.
        assert matriculas_na_ordem == ["sem-hora", "com-hora", "sem-data-nem-hora"], matriculas_na_ordem
        print("OK: registro sem hora (e sem data) mantém a posição física, sem ir para o final")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teste_aba_revisao_tambem_preserva_a_ordem_fisica():
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida_revisao_ordem.xlsx")
        registros = [
            _registro("REVISAO", matricula="rev1", data="15/08/2026", hora="16:10"),
            _registro("CONFIRMADO", matricula="ok1", data="14/08/2026", hora="08:15"),
            _registro("REVISAO", matricula="rev2", data="14/08/2026", hora="07:42"),
        ]
        export_to_xlsx(registros, caminho)

        wb = openpyxl.load_workbook(caminho)
        ws = wb["Revisão"]
        matriculas_na_ordem = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        assert matriculas_na_ordem == ["rev1", "rev2"], matriculas_na_ordem
        print("OK: aba Revisão preserva a ordem física (rev1 antes de rev2)")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teste_sem_problematicos():
    tmp = tempfile.mkdtemp()
    try:
        caminho = os.path.join(tmp, "saida4.xlsx")
        export_to_xlsx([_registro("CONFIRMADO")], caminho)
        wb = openpyxl.load_workbook(caminho)
        ws = wb["Revisão"]
        assert ws.max_row == 1  # só o cabeçalho, nenhum problema
        print("OK: sem registros problemáticos, aba Revisão fica só com cabeçalho")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    testes = [v for k, v in list(globals().items()) if k.startswith("teste_")]
    falhas = 0
    for t in testes:
        try:
            t()
        except AssertionError as exc:
            falhas += 1
            print(f"FALHOU: {t.__name__}: {exc}")
    print("=" * 50)
    print(f"TODOS OS {len(testes)} TESTES PASSARAM." if not falhas else f"{len(testes)-falhas}/{len(testes)} passaram")
    if falhas:
        sys.exit(1)
