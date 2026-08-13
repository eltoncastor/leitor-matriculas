"""
teste/teste_seguranca_lote.py (Fase 8 — segurança operacional do lote)

Simula uma falha INESPERADA no meio de um lote (não um arquivo de foto
ruim -- isso já é coberto por teste_lote_operacional.py -- mas uma
exceção que escapa do próprio consumidor da fila no thread principal,
ex.: um dado malformado). Confirma que:

    1. o polling da fila NÃO trava (self.after volta a ser reagendado);
    2. o lote continua processando as páginas seguintes normalmente;
    3. os resultados das páginas anteriores/posteriores à falha
       continuam disponíveis em memória;
    4. a cópia de segurança automática (_autosave_lote) existe em disco
       e reflete corretamente o que foi processado;
    5. "Gerar planilha" ainda funciona com o que sobrou.

PaddleOCR MOCKADO. Uso:
    python teste\\teste_seguranca_lote.py
"""
import os
import sys
import time
import shutil
import tempfile
import threading

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import numpy as np  # noqa: E402
import cv2  # noqa: E402
from unittest.mock import patch, MagicMock  # noqa: E402

from leitor_matriculas.ui.app import App  # noqa: E402
from leitor_matriculas.ocr.engine import OCRResult  # noqa: E402

TOTAL_PAGINAS = 10
PAGINA_QUE_FALHA = 5  # falha simulada no CONSUMIDOR da fila (thread principal), não no arquivo

COL_DATA, COL_HORA, COL_NOME, COL_MAT, COL_SETOR, COL_MOT, COL_GESTOR = 60, 160, 280, 450, 560, 700, 880


def _r(t, x1, y1, x2, y2, c=0.95):
    return OCRResult(texto_original=t, confianca=c, box=[x1, y1, x2, y2])


def _pagina_ocr_com_8_registros():
    elementos = [
        _r("DATA", COL_DATA - 20, 10, COL_DATA + 20, 30), _r("HORA", COL_HORA - 20, 10, COL_HORA + 20, 30),
        _r("NOME", COL_NOME - 25, 10, COL_NOME + 25, 30), _r("MATRÍCULA", COL_MAT - 35, 10, COL_MAT + 35, 30),
        _r("SETOR", COL_SETOR - 25, 10, COL_SETOR + 25, 30), _r("MOTIVO", COL_MOT - 25, 10, COL_MOT + 25, 30),
        _r("RESPONSÁVEL", COL_GESTOR - 40, 10, COL_GESTOR + 40, 30),
    ]
    for i in range(8):
        y1, y2 = 40 + i * 30, 60 + i * 30
        elementos += [
            _r("23.04.2026", COL_DATA - 15, y1, COL_DATA + 15, y2),
            _r("11:05", COL_HORA - 15, y1, COL_HORA + 15, y2),
            _r("Fulano", COL_NOME - 25, y1, COL_NOME + 25, y2),
            _r(f"{28000 + i}", COL_MAT - 20, y1, COL_MAT + 20, y2),
            _r("TI", COL_SETOR - 20, y1, COL_SETOR + 20, y2),
            _r("RH", COL_MOT - 25, y1, COL_MOT + 25, y2),
            _r("Gestor X", COL_GESTOR - 30, y1, COL_GESTOR + 30, y2),
        ]
    return elementos


falhas = []


def checar(cond, msg):
    print(("  OK    " if cond else "  FALHA ") + msg)
    if not cond:
        falhas.append(msg)


tmp = tempfile.mkdtemp()
tmp_saida = tempfile.mkdtemp()  # pasta de autosave isolada, não a saida/ real do projeto
try:
    caminhos = []
    for i in range(TOTAL_PAGINAS):
        caminho = os.path.join(tmp, f"folha_{i:03d}.jpg")
        cv2.imwrite(caminho, np.full((120, 900, 3), 255, dtype=np.uint8))
        caminhos.append(caminho)

    original_adicionar = App._adicionar_registros

    def _adicionar_registros_com_falha(self, numero_pagina, registros):
        if numero_pagina == PAGINA_QUE_FALHA:
            # Simula uma exceção INESPERADA no consumidor da fila (thread
            # principal) -- não um arquivo de imagem ruim (já coberto em
            # teste_lote_operacional.py), mas algo que escaparia de
            # _processar_item antes da Fase 8.
            raise RuntimeError("falha simulada no meio do lote (item malformado)")
        return original_adicionar(self, numero_pagina, registros)

    with patch("leitor_matriculas.ui.app.Messagebox.show_error"), \
         patch("leitor_matriculas.ui.app.Messagebox.show_warning"), \
         patch("leitor_matriculas.ui.app.Messagebox.show_info"), \
         patch.object(App, "_adicionar_registros", _adicionar_registros_com_falha), \
         patch.object(App, "_pasta_saida_padrao", staticmethod(lambda: tmp_saida)):
        app = App()
        fake_engine = MagicMock()
        fake_engine.recognize.return_value = _pagina_ocr_com_8_registros()
        app._ocr_engine = fake_engine

        print(f"=== Lote de {TOTAL_PAGINAS} páginas, falha simulada na página {PAGINA_QUE_FALHA} ===")
        app._iniciar_processamento(f"Processando 0/{TOTAL_PAGINAS} imagens ...")
        threading.Thread(target=app._worker_imagens, args=(caminhos,), daemon=True).start()
        app.after(100, app._verificar_fila)

        t0 = time.time()
        while app._processando:
            app.update()
            time.sleep(0.02)
            if time.time() - t0 > 60:
                raise AssertionError("lote travou -- polling da fila não se recuperou da falha simulada")

        # -------------------------------------------------------------
        # 1) O polling não travou: o lote chegou até o fim.
        # -------------------------------------------------------------
        checar(app._processando is False, "processamento chegou ao fim (polling não travou com a falha)")

        # -------------------------------------------------------------
        # 2) Todas as 10 páginas foram contabilizadas, mesmo a que falhou.
        # -------------------------------------------------------------
        checar(app._paginas_processadas == TOTAL_PAGINAS,
               f"todas as {TOTAL_PAGINAS} páginas contabilizadas (obtido: {app._paginas_processadas})")

        # -------------------------------------------------------------
        # 3) Páginas ANTES e DEPOIS da falha continuam disponíveis (a
        #    falha isolada não apagou nada que já tinha sido processado
        #    nem impediu as páginas seguintes).
        # -------------------------------------------------------------
        paginas_com_registros = sorted({r["pagina_origem"] for r in app._registros_exportacao})
        esperado = [p for p in range(1, TOTAL_PAGINAS + 1) if p != PAGINA_QUE_FALHA]
        checar(paginas_com_registros == esperado,
               f"páginas com registro: {paginas_com_registros} (esperado, sem a {PAGINA_QUE_FALHA}: {esperado})")
        checar(len(app._registros_exportacao) == len(esperado) * 8,
               f"{len(esperado) * 8} registros das páginas bem-sucedidas presentes "
               f"(obtido: {len(app._registros_exportacao)})")

        # -------------------------------------------------------------
        # 4) Ordem física preservada (a falha no meio não embaralhou
        #    nada -- as páginas continuam na ordem em que vieram).
        # -------------------------------------------------------------
        ordem_paginas_na_lista = [r["pagina_origem"] for r in app._registros_exportacao]
        checar(ordem_paginas_na_lista == sorted(ordem_paginas_na_lista),
               "ordem física preservada mesmo com falha no meio do lote")

        # -------------------------------------------------------------
        # 5) Autosave em disco reflete o estado real (Fase 8).
        # -------------------------------------------------------------
        caminho_autosave = os.path.join(tmp_saida, "Liberacoes_autosave.xlsx")
        checar(os.path.isfile(caminho_autosave), "cópia de segurança automática foi gravada em disco")
        import openpyxl
        wb_auto = openpyxl.load_workbook(caminho_autosave)
        checar(wb_auto.sheetnames == ["Liberações", "Revisão", "Resumo"],
               f"autosave tem as 3 abas corretas: {wb_auto.sheetnames}")
        checar(wb_auto["Liberações"].max_row - 1 == len(esperado) * 8,
               f"autosave tem {wb_auto['Liberações'].max_row - 1} linhas "
               f"(esperado {len(esperado) * 8}) -- reflete o processado até a última página")

        # -------------------------------------------------------------
        # 6) "Gerar planilha" final ainda funciona com o que sobrou.
        # -------------------------------------------------------------
        saida_final = os.path.join(tmp, "final.xlsx")
        with patch("leitor_matriculas.ui.app.filedialog.asksaveasfilename", return_value=saida_final):
            app._on_salvar()
        checar(os.path.isfile(saida_final), "XLSX final gerado com sucesso após a falha no meio do lote")
        wb_final = openpyxl.load_workbook(saida_final)
        checar(wb_final["Liberações"].max_row - 1 == len(esperado) * 8,
               f"XLSX final tem {wb_final['Liberações'].max_row - 1} linhas (esperado {len(esperado) * 8})")

        app.destroy()
finally:
    shutil.rmtree(tmp, ignore_errors=True)
    shutil.rmtree(tmp_saida, ignore_errors=True)

print()
print("=" * 60)
if not falhas:
    print("TESTE DE SEGURANÇA DO LOTE (FALHA NO MEIO): TUDO OK")
else:
    print(f"TESTE DE SEGURANÇA DO LOTE: {len(falhas)} FALHA(S)")
    sys.exit(1)
