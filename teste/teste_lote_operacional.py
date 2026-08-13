"""
teste/teste_lote_operacional.py (Fase 7 — preparação para operação real)

Simula o cenário de amanhã: ~50 fotos selecionadas de uma vez (seleção
múltipla de imagens, PROBLEMA da Fase 7), com algumas fotos CORROMPIDAS
de propósito espalhadas no meio do lote. PaddleOCR MOCKADO (o motor de
OCR em si não muda nesta fase) -- o que este teste verifica é a
ROBUSTEZ OPERACIONAL do lote: nenhuma página perdida, arquivo corrompido
isolado (não derruba o lote), progresso rastreável, contagem final
consistente, XLSX gerado com todas as páginas presentes.

Uso:
    python teste\\teste_lote_operacional.py
"""
import os
import sys
import time
import shutil
import tempfile
import threading

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import numpy as np  # noqa: E402
import cv2  # noqa: E402
from unittest.mock import patch, MagicMock  # noqa: E402

from leitor_matriculas.ui.app import App  # noqa: E402
from leitor_matriculas.ocr.engine import OCRResult  # noqa: E402

TOTAL_PAGINAS = 50
INDICES_CORROMPIDOS = {5, 17, 33}  # 3 fotos "ruins" espalhadas no lote

COL_DATA, COL_HORA, COL_NOME, COL_MAT, COL_SETOR, COL_MOT, COL_GESTOR = 60, 160, 280, 450, 560, 700, 880


def _r(t, x1, y1, x2, y2, c=0.95):
    return OCRResult(texto_original=t, confianca=c, box=[x1, y1, x2, y2])


def _cabecalho(y1=10, y2=30):
    return [
        _r("DATA", COL_DATA - 20, y1, COL_DATA + 20, y2), _r("HORA", COL_HORA - 20, y1, COL_HORA + 20, y2),
        _r("NOME", COL_NOME - 25, y1, COL_NOME + 25, y2), _r("MATRÍCULA", COL_MAT - 35, y1, COL_MAT + 35, y2),
        _r("SETOR", COL_SETOR - 25, y1, COL_SETOR + 25, y2), _r("MOTIVO", COL_MOT - 25, y1, COL_MOT + 25, y2),
        _r("RESPONSÁVEL", COL_GESTOR - 40, y1, COL_GESTOR + 40, y2),
    ]


def _pagina_ocr_com_8_registros():
    """OCR simulado de uma folha normal: cabeçalho + 8 liberações completas."""
    elementos = list(_cabecalho())
    for i in range(8):
        y1, y2 = 40 + i * 30, 60 + i * 30
        elementos += [
            _r("23.04.2026", COL_DATA - 15, y1, COL_DATA + 15, y2),
            _r("11:05", COL_HORA - 15, y1, COL_HORA + 15, y2),
            _r("Fulano", COL_NOME - 25, y1, COL_NOME + 25, y2),
            _r(f"{28000 + i}", COL_MAT - 20, y1, COL_MAT + 20, y2),
            _r("TI", COL_SETOR - 20, y1, COL_SETOR + 20, y2),
            _r("RH", COL_MOT - 25, y1, COL_MOT + 25, y2),
            _r("Gestor X", COL_GESTOR - 30, y1, COL_GESTOR + 30, y2),
        ]
    return elementos


def _rodar_ate_terminar(app, timeout_s=60):
    t0 = time.time()
    while app._processando:
        app.update()
        time.sleep(0.02)
        if time.time() - t0 > timeout_s:
            raise AssertionError(f"lote não terminou em {timeout_s}s -- ficaria preso em 'Processando...'")


falhas = []


def checar(cond, msg):
    print(("  OK    " if cond else "  FALHA ") + msg)
    if not cond:
        falhas.append(msg)


tmp = tempfile.mkdtemp()
try:
    caminhos = []
    for i in range(TOTAL_PAGINAS):
        caminho = os.path.join(tmp, f"folha_{i:03d}.jpg")
        if i in INDICES_CORROMPIDOS:
            # Foto "corrompida": arquivo existe mas não é uma imagem válida
            # -- exatamente o tipo de arquivo ruim que pode aparecer entre
            # 50 fotos reais tiradas às pressas (arquivo truncado, 0 bytes,
            # formato errado).
            with open(caminho, "wb") as f:
                f.write(b"isto nao e um jpeg valido")
        else:
            cv2.imwrite(caminho, np.full((120, 900, 3), 255, dtype=np.uint8))
        caminhos.append(caminho)

    with patch("leitor_matriculas.ui.app.Messagebox.show_error"), \
         patch("leitor_matriculas.ui.app.Messagebox.show_warning"), \
         patch("leitor_matriculas.ui.app.Messagebox.show_info"):
        app = App()
        fake_engine = MagicMock()
        fake_engine.recognize.return_value = _pagina_ocr_com_8_registros()
        app._ocr_engine = fake_engine

        print(f"=== Lote de {TOTAL_PAGINAS} imagens, {len(INDICES_CORROMPIDOS)} corrompidas ===")
        app._iniciar_processamento(f"Processando 0/{TOTAL_PAGINAS} imagens ...")
        threading.Thread(target=app._worker_imagens, args=(caminhos,), daemon=True).start()
        app.after(100, app._verificar_fila)
        _rodar_ate_terminar(app)

        # -------------------------------------------------------------
        # 1) Nenhuma página perdida: todas as 50 contabilizadas.
        # -------------------------------------------------------------
        checar(app._paginas_processadas == TOTAL_PAGINAS,
               f"todas as {TOTAL_PAGINAS} páginas contabilizadas (obtido: {app._paginas_processadas})")
        checar(app._total_paginas_lote == TOTAL_PAGINAS, "total do lote conhecido e correto (progresso rastreável)")
        checar(app._paginas_processadas_lote == TOTAL_PAGINAS, "contador de progresso da leva bate com o total")

        # -------------------------------------------------------------
        # 2) Isolamento: as 3 fotos corrompidas viram ERRO, sem derrubar
        #    as outras 47.
        # -------------------------------------------------------------
        checar(app._paginas_com_erro == len(INDICES_CORROMPIDOS),
               f"{len(INDICES_CORROMPIDOS)} páginas com erro isoladas (obtido: {app._paginas_com_erro})")
        paginas_com_erro_relatadas = {e["pagina"] for e in app._erros_paginas}
        esperado_erro = {i + 1 for i in INDICES_CORROMPIDOS}  # numero_pagina é 1-based, na ordem da lista
        checar(paginas_com_erro_relatadas == esperado_erro,
               f"números de página com erro batem com os arquivos corrompidos: {sorted(paginas_com_erro_relatadas)}")

        # -------------------------------------------------------------
        # 3) Rastreabilidade: todo número de página 1..50 aparece em
        #    algum registro exportado (nenhuma página "sumiu" da saída).
        # -------------------------------------------------------------
        paginas_na_exportacao = {r["pagina_origem"] for r in app._registros_exportacao}
        checar(paginas_na_exportacao == set(range(1, TOTAL_PAGINAS + 1)),
               "todo número de página 1..50 aparece em pelo menos um registro exportado")

        # -------------------------------------------------------------
        # 4) Contagem final: 47 páginas boas x 8 registros + 3 ERRO = 379.
        # -------------------------------------------------------------
        esperado_registros = (TOTAL_PAGINAS - len(INDICES_CORROMPIDOS)) * 8 + len(INDICES_CORROMPIDOS)
        checar(len(app._registros_exportacao) == esperado_registros,
               f"contagem final de registros bate ({len(app._registros_exportacao)} == {esperado_registros})")

        erro_rows = [r for r in app._registros_exportacao if r["status"] == "ERRO"]
        checar(len(erro_rows) == len(INDICES_CORROMPIDOS), "linhas ERRO correspondem às fotos corrompidas")
        checar(all(not r["matricula"] for r in erro_rows), "linha ERRO não inventa matrícula nenhuma")

        # -------------------------------------------------------------
        # 5) UI recuperada ao final (não fica presa em Processando...).
        # -------------------------------------------------------------
        checar(app._processando is False, "processamento encerrado corretamente")
        checar(str(app.btn_salvar.cget("state")) == "normal", "botão Gerar planilha habilitado ao final")

        # -------------------------------------------------------------
        # 6) XLSX final: gera e confere linha a linha.
        # -------------------------------------------------------------
        saida_xlsx = os.path.join(tmp, "lote.xlsx")
        with patch("leitor_matriculas.ui.app.filedialog.asksaveasfilename", return_value=saida_xlsx):
            app._on_salvar()
        checar(os.path.isfile(saida_xlsx), "XLSX do lote foi gerado")

        import openpyxl
        wb = openpyxl.load_workbook(saida_xlsx)
        checar(wb.sheetnames == ["Liberações", "Revisão", "Resumo"], f"abas corretas: {wb.sheetnames}")
        ws_lib = wb["Liberações"]
        checar(ws_lib.max_row - 1 == esperado_registros,
               f"XLSX tem {ws_lib.max_row - 1} linhas de dado (esperado {esperado_registros})")
        # NOTA: o DataManager aqui é o real (dados/ da máquina local), não
        # mockado -- as matrículas sintéticas 28000-28007 podem ou não
        # existir na base real, então parte das 47 páginas "boas" pode
        # legitimamente cair em REVISAO (não é o alvo deste teste; ver
        # teste_validacao.py para as regras de classificação em si). O
        # invariante que interessa aqui é estrutural: a aba Revisão
        # contém exatamente as linhas REVISAO+ERRO, nem mais nem menos.
        problematicos_esperados = [r for r in app._registros_exportacao if r["status"] in ("REVISAO", "ERRO")]
        ws_rev = wb["Revisão"]
        checar(ws_rev.max_row - 1 == len(problematicos_esperados),
               f"aba Revisão contém exatamente as linhas REVISAO+ERRO "
               f"({ws_rev.max_row - 1} == {len(problematicos_esperados)})")
        checar(len(erro_rows) == len(INDICES_CORROMPIDOS),
               f"dessas, {len(INDICES_CORROMPIDOS)} são ERRO das fotos corrompidas (obtido: {len(erro_rows)})")

        # Ordem física preservada mesmo em lote grande (Fase 2, não deve
        # regredir): a primeira página no arquivo continua a primeira na
        # planilha.
        primeira_pagina_xlsx = ws_lib.cell(row=2, column=9).value  # coluna "Página"
        checar(primeira_pagina_xlsx == 1, f"primeira linha da planilha é da página 1 (obtido: {primeira_pagina_xlsx})")

        app.destroy()
finally:
    shutil.rmtree(tmp, ignore_errors=True)

print()
print("=" * 60)
if not falhas:
    print("TESTE DE LOTE OPERACIONAL (50 PÁGINAS): TUDO OK")
else:
    print(f"TESTE DE LOTE OPERACIONAL: {len(falhas)} FALHA(S)")
    sys.exit(1)
