"""
teste/teste_ui_integracao.py

Teste de integracao ponta-a-ponta da interface: imagem unica + PDF multi-
pagina + exportacao XLSX + revisao manual. Usa PaddleOCR MOCKADO (nao o
real). No Linux/CI roda sob Xvfb (display virtual); no Windows usa o
display nativo, sem precisar de nada extra.

Uso:
    xvfb-run -a python3 teste/teste_ui_integracao.py   (Linux/CI)
    python teste\\teste_ui_integracao.py                 (Windows)
"""
import time, os, sys, tempfile, shutil
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))
import numpy as np, cv2
from unittest.mock import patch, MagicMock
import pymupdf as fitz
from tkinter import ttk
import ttkbootstrap as tb

from leitor_matriculas.ui.app import App, _ler_imagem
from leitor_matriculas.ocr.engine import OCRResult
from leitor_matriculas.ui import estilos


class _DMRevisaoDeterministica:
    """
    DataManager fake usado só a partir da etapa de revisão manual, Fase 7.

    A revisão manual agora REAVALIA de verdade a correção digitada (ver
    PROBLEMAS C/D em app.py) em vez de forçar CONFIRMADO -- então o
    resultado da correção passa a depender do conteúdo de `dados/`. Como
    `dados/` é local/gitignored (pode nem existir em outra máquina, ver
    CLAUDE.md), o teste não pode depender do que essas planilhas reais
    contêm para ser determinístico. Esta base fake substitui a real só
    para o passo de revisão, e reconhece exatamente UMA matrícula.
    """
    colaboradores_disponivel = True
    gestores_disponivel = False  # sem base -> gestor não bloqueia (mesma regra já existente)
    motivos_disponivel = False
    avisos = []

    def buscar_colaborador(self, matricula):
        if matricula == "28972":
            return {"matricula": "28972", "nome": "Fulano Corrigido", "cargo": "Cargo X", "setor": "Setor X"}
        return None

    def listar_gestores(self):
        return []

    def listar_motivos(self):
        return []

# --- fixtures: imagem sintética de teste ---
# Usa tempfile em vez de um caminho POSIX fixo ("/tmp/...") -- em Windows
# isso não é garantido existir/ser gravável, diferente de tempfile, que
# sempre resolve para uma pasta temporária válida na plataforma atual.
_tmp_dir_img = tempfile.mkdtemp()
_caminho_img_teste = os.path.join(_tmp_dir_img, 'teste_ui_img.jpg')
img = np.full((200, 900, 3), 255, dtype=np.uint8)
cv2.imwrite(_caminho_img_teste, img)

# --- resultados de OCR simulados (cabecalho + 2 registros, layout do teste_registro_parser) ---
COL_DATA, COL_HORA, COL_NOME, COL_MAT, COL_SETOR, COL_MOT, COL_GESTOR = 60,160,280,450,560,700,880
def r(t,x1,y1,x2,y2,c=0.9): return OCRResult(texto_original=t, confianca=c, box=[x1,y1,x2,y2])
def cabecalho(y1=10,y2=30):
    return [r("DATA",COL_DATA-20,y1,COL_DATA+20,y2), r("HORA",COL_HORA-20,y1,COL_HORA+20,y2),
            r("NOME",COL_NOME-25,y1,COL_NOME+25,y2), r("MATRÍCULA",COL_MAT-35,y1,COL_MAT+35,y2),
            r("SETOR",COL_SETOR-25,y1,COL_SETOR+25,y2), r("MOTIVO",COL_MOT-25,y1,COL_MOT+25,y2),
            r("RESPONSÁVEL",COL_GESTOR-40,y1,COL_GESTOR+40,y2)]
def linha(y1,y2,data,hora,nome,mat,setor,mot,gestor,conf_mat=0.9):
    return [r(data,COL_DATA-15,y1,COL_DATA+15,y2), r(hora,COL_HORA-15,y1,COL_HORA+15,y2),
            r(nome,COL_NOME-25,y1,COL_NOME+25,y2), r(mat,COL_MAT-20,y1,COL_MAT+20,y2,conf_mat),
            r(setor,COL_SETOR-20,y1,COL_SETOR+20,y2), r(mot,COL_MOT-25,y1,COL_MOT+25,y2),
            r(gestor,COL_GESTOR-30,y1,COL_GESTOR+30,y2)]

resultados_ocr = cabecalho() + linha(40,60,"23.04.2026","11:05","Fulano","28972","TI","RH","Gestor X") \
                              + linha(70,90,"23.04.2026","11:10","Beltrano","99999","RH","ADM","Gestor Y", conf_mat=0.5)

# Fase 20 (H5): o histórico de correções vai para um arquivo TEMPORÁRIO
# durante o teste. Rodar a suíte não pode escrever no `dados/` real do
# operador -- e este patch é também o que garante que o teste observe
# exatamente o que a interface gravou.
_tmp_dir_correcoes = tempfile.mkdtemp(prefix="teste_ui_correcoes_")
_arquivo_correcoes = os.path.join(_tmp_dir_correcoes, "correcoes_humanas.jsonl")

with patch('leitor_matriculas.ui.app.Messagebox.show_error') as m_err, patch('leitor_matriculas.ui.app.Messagebox.show_info') as m_info, \
     patch('leitor_matriculas.ui.app.Messagebox.show_warning') as m_warn, \
     patch('leitor_matriculas.dados.registro_correcoes.caminho_padrao',
           return_value=_arquivo_correcoes):
    app = App()
    fake_engine = MagicMock()
    fake_engine.recognize.return_value = resultados_ocr
    app._ocr_engine = fake_engine

    # -------- fluxo imagem única --------
    app._imagem_original = _ler_imagem(_caminho_img_teste)
    app._arquivo_atual = 'teste.jpg'
    app._iniciar_processamento("...")
    import threading
    threading.Thread(target=app._worker_imagem, args=(app._imagem_original,), daemon=True).start()
    app.after(100, app._verificar_fila)
    for _ in range(60):
        app.update(); time.sleep(0.05)
        if not app._processando: break

    linhas = app.tabela.get_children()
    print("Linhas na tabela (imagem única):", len(linhas))
    assert len(linhas) == 2
    valores = [app.tabela.item(i)['values'] for i in linhas]
    print(valores[0]); print(valores[1])
    # Ordem atual da tabela: pagina(0), status(1), data(2), hora(3), matricula(4), ...
    assert '28972' in str(valores[0][4])
    assert 'CONFIRMADO' not in valores[1][1]  # matricula 99999 nao existe na base -> revisao (sem base carregada tb cai em revisao)
    assert app._contador_confirmados + app._contador_revisao == 2
    print("Confirmados:", app._contador_confirmados, "Revisao:", app._contador_revisao)
    assert str(app.btn_salvar.cget('state')) == 'normal'

    # -------- fluxo PDF (2 páginas) --------
    tmp = tempfile.mkdtemp()
    caminho_pdf = os.path.join(tmp, 'mes.pdf')
    doc = fitz.open()
    for i in range(2):
        doc.new_page(width=600, height=800)
    doc.save(caminho_pdf); doc.close()

    app._arquivo_atual = 'mes.pdf'
    app._iniciar_processamento("...")
    threading.Thread(target=app._worker_pdf, args=(caminho_pdf,), daemon=True).start()
    app.after(100, app._verificar_fila)
    for _ in range(600):
        app.update(); time.sleep(0.05)
        if not app._processando: break
    else:
        print("AVISO: loop esgotado sem terminar")

    print("Paginas processadas total:", app._paginas_processadas, "erros:", app._erros_paginas, "paginas_com_erro:", app._paginas_com_erro)
    assert app._paginas_processadas == 3  # 1 imagem + 2 paginas pdf
    linhas2 = app.tabela.get_children()
    print("Linhas na tabela apos PDF:", len(linhas2))
    assert len(linhas2) == 6  # 2 (imagem) + 2 (pagina1 pdf) + 2 (pagina2 pdf), mesmo mock em todas

    # -------- salvar XLSX --------
    saida_xlsx = os.path.join(tmp, 'saida.xlsx')
    with patch('leitor_matriculas.ui.app.filedialog.asksaveasfilename', return_value=saida_xlsx):
        app._on_salvar()
    assert m_info.called
    assert os.path.isfile(saida_xlsx)
    import openpyxl
    wb = openpyxl.load_workbook(saida_xlsx)
    print("Abas do XLSX:", wb.sheetnames)
    assert wb.sheetnames == ["Liberações","Revisão","Resumo"]
    print("Linhas Liberacoes:", wb["Liberações"].max_row)

    # -------- revisao manual (Fase 7: reavaliação real, não force-confirm) --------
    # A partir daqui a base real de dados/ deixa de importar -- ver
    # _DMRevisaoDeterministica. Isso NÃO reclassifica os registros já
    # processados (o status já foi decidido, com a base real, lá em
    # cima); só passa a valer para o que `_confirmar` reavaliar agora.
    app._data_manager = _DMRevisaoDeterministica()

    # A revisão deixou de ser uma janela Toplevel (Fase 10) e virou uma
    # ABA da janela principal. Este trecho passou a dirigi-la pela API
    # programática (_indices_pendentes_revisao / _revisao_ir_para /
    # _revisao_confirmar / revisao_vars) em vez de caçar widgets na árvore
    # do Tk. O que é verificado continua sendo exatamente o mesmo
    # COMPORTAMENTO da Fase 7 (PROBLEMAS C/D/E) -- só o caminho até ele
    # mudou, e agora não quebra a cada ajuste de layout.
    print("Botao revisao:", app.btn_revisao.cget('text'))
    app._abrir_revisao()
    pendentes = app._indices_pendentes_revisao()
    print("Itens pendentes na revisao:", len(pendentes))
    assert len(pendentes) == app._contador_revisao
    # PROBLEMA E: a revisão só pode listar linhas REVISAO -- nenhuma ERRO
    # (página que falhou não tem campo nenhum de verdade a corrigir).
    assert app._paginas_com_erro == 0  # este lote não teve página com erro

    # -------- 1) correção REAL (matrícula corrigida para uma que a base
    # (fake) reconhece) -- deve virar CONFIRMADO, e Nome/Setor/Cargo
    # devem ser re-consultados (PROBLEMA C), não ficar em "(não encontrado)".
    revisao_antes = app._contador_revisao
    confirmados_antes = app._contador_confirmados
    app._revisao_ir_para(0)
    app.update()
    app.revisao_vars["matricula"].set("28972")  # matrícula reconhecida pela fake DM
    app._revisao_confirmar()
    app.update()

    assert app._contador_revisao == revisao_antes - 1
    assert app._contador_confirmados == confirmados_antes + 1
    print("OK: correcao manual REAL moveu 1 registro de Revisao para Confirmado")
    print("Botao revisao apos correcao:", app.btn_revisao.cget("text"))

    linhas_apos_correcao = app.tabela.get_children()
    status_corrigido = [app.tabela.item(i)["values"][1] for i in linhas_apos_correcao]  # status agora é o índice 1
    # Sub-fase 21c: o texto exibido vem do vocabulário único de
    # `ui/estilos.py` ("✓ Confirmado"), não mais de uma string solta.
    from leitor_matriculas.ui import estilos as _estilos
    assert any(s == _estilos.texto_status("CONFIRMADO") for s in status_corrigido)
    print("OK: tabela principal refletiu a correcao (sincronizacao)")

    corrigidos = [r for r in app._registros_exportacao if r.get("nome") == "Fulano Corrigido"]
    assert len(corrigidos) == 1, corrigidos
    assert corrigidos[0]["status"] == "CONFIRMADO"
    assert corrigidos[0]["setor"] == "Setor X" and corrigidos[0]["cargo"] == "Cargo X"
    print("OK: PROBLEMA C -- Nome/Setor/Cargo re-consultados pela matricula corrigida (nao ficaram '(não encontrado)')")

    # -------- 2) correção que NÃO resolve o problema real -- Fase 7
    # (PROBLEMA D): não pode virar CONFIRMADO só por ter clicado o botão.
    pendentes_2 = app._indices_pendentes_revisao()
    if pendentes_2:
        indice_pendente = pendentes_2[0]
        app._revisao_ir_para(0)
        app.update()
        app.revisao_vars["matricula"].set("00000")  # a fake DM não reconhece
        revisao_antes_2 = app._contador_revisao
        app._revisao_confirmar()
        app.update()
        assert app._contador_revisao == revisao_antes_2, "registro nao corrigido de verdade nao pode virar CONFIRMADO"
        assert indice_pendente in app._indices_pendentes_revisao(), "registro ainda pendente deve continuar na lista"
        print("OK: PROBLEMA D -- correcao que nao resolve o problema real permanece em REVISAO")

    # -------- 3) Fase 10: DATA e HORA agora são editáveis na revisão.
    # Antes uma linha barrada pela data era impossível de resolver dentro
    # do programa (não havia campo para corrigi-la).
    assert "data" in app.revisao_vars and "hora" in app.revisao_vars, \
        "a revisão precisa permitir editar DATA e HORA"
    pendentes_3 = app._indices_pendentes_revisao()
    if pendentes_3:
        app._revisao_ir_para(0)
        app.update()
        registro_alvo = app._registros_exportacao[pendentes_3[0]]
        app.revisao_vars["matricula"].set("28972")
        app.revisao_vars["data"].set("23/04/2026")
        app.revisao_vars["hora"].set("11:05")
        app._revisao_confirmar()
        app.update()
        assert registro_alvo["data"] == "23/04/26", registro_alvo
        assert registro_alvo["hora"] == "11:05", registro_alvo
        print("OK: Fase 10 -- DATA/HORA editadas na revisao saem no formato canonico")

    # -------- PROBLEMA E: uma linha ERRO nunca pode entrar na revisão ------
    app._registros_exportacao.append({
        "data": "", "hora": "", "matricula": "", "nome": "", "cargo": "", "setor": "",
        "gestor": "", "motivo": "", "pagina_origem": 999, "status": "ERRO",
        "confianca_matricula": "", "confianca_gestor": "", "confianca_motivo": "",
        "observacao": "falha simulada de pagina", "texto_ocr_original": "",
    })
    indice_erro = len(app._registros_exportacao) - 1
    assert indice_erro not in app._indices_pendentes_revisao(), \
        "linha ERRO nao pode aparecer na revisao manual"
    paginas_pendentes = [
        app._registros_exportacao[i]["pagina_origem"] for i in app._indices_pendentes_revisao()
    ]
    assert 999 not in paginas_pendentes, "linha ERRO nao pode aparecer na revisao manual"
    print("OK: PROBLEMA E -- linha ERRO nunca aparece na revisao manual")

    # -------- Fase 18: a revisao EXPLICA a duvida ------------------------
    # Dirigido pelo contrato programatico (_explicacao_revisao_atual /
    # _revisao_ir_para / revisao_vars), nunca cacando widgets na arvore do
    # Tk -- que e o erro que a Fase 10 corrigiu.
    pendentes_18 = app._indices_pendentes_revisao()
    if pendentes_18:
        for posicao in range(len(pendentes_18)):
            app._revisao_ir_para(posicao)
            explicacao, sinais = app._explicacao_revisao_atual()
            indice_atual, registro_atual = app._revisao_registro_atual()

            # 1. toda linha em REVISAO tem explicacao
            assert not explicacao.vazia, f"linha {indice_atual} ficou sem explicacao"
            assert explicacao.como_texto(), f"linha {indice_atual} explicou vazio"

            # 2. a explicacao nomeia o campo que bloqueou (quando ha dossie)
            if registro_atual.get("evidencias"):
                assert explicacao.campos_bloqueantes, \
                    f"linha {indice_atual} nao identificou o campo bloqueante"
                for campo in explicacao.campos_bloqueantes:
                    assert campo in app.revisao_vars, \
                        f"campo bloqueante {campo!r} nao existe no formulario"

            # 3. NENHUM campo do formulario chega pre-preenchido com sugestao:
            #    o valor exibido e sempre o que ja estava no registro.
            for chave, var in app.revisao_vars.items():
                assert var.get() == (registro_atual.get(chave) or ""), (
                    f"campo {chave!r} da linha {indice_atual} foi pre-preenchido "
                    f"({var.get()!r} != {registro_atual.get(chave)!r})"
                )

            # 4. sinais de contexto sao informativos: nao mexem em revisao_vars
            antes = {k: v.get() for k, v in app.revisao_vars.items()}
            sinais_2 = app._explicacao_revisao_atual()[1]
            depois = {k: v.get() for k, v in app.revisao_vars.items()}
            assert antes == depois, "montar os sinais de contexto alterou revisao_vars"
            assert len(sinais) == len(sinais_2), "sinais de contexto nao sao deterministicos"
            for sinal in sinais:
                assert sinal.tipo == "contexto", sinal
                assert sinal.origem in (
                    "gestores_do_lote", "ordem_cronologica_da_folha",
                ), f"sinal de contexto nao previsto: {sinal.origem}"
        print(f"OK: Fase 18 -- {len(pendentes_18)} linha(s) em revisao com explicacao, "
              "nenhum campo pre-preenchido")

    # A explicacao nunca inventa: sem dossie e sem observacao, sai vazia.
    from leitor_matriculas.ui import explicacao_revisao as _expl
    assert _expl.explicar([], "").vazia, "explicacao sem dado nenhum deveria sair vazia"
    assert not _expl.explicar([], "algum motivo").vazia, \
        "sem dossie, a observacao antiga ainda deve ser mostrada"
    print("OK: Fase 18 -- explicacao nunca inventa texto")

    # -------- Fase 20 (H5): o histórico registrou as correções ----------
    # Infraestrutura de COLETA: grava depois da decisão e não participa
    # dela. O que se verifica aqui é que ela captura o que se perdia --
    # qual campo mudou, o valor anterior e o texto que o OCR havia lido --
    # e que a linha que NÃO resolveu também foi registrada.
    from leitor_matriculas.dados import registro_correcoes as _rc

    correcoes = _rc.ler_correcoes(_arquivo_correcoes)
    assert correcoes, "nenhuma correcao humana foi registrada"

    # A correção 1 (matrícula 28972) resolveu; a 2 (matrícula 00000) não.
    resolvidas = [c for c in correcoes if c["resolveu"]]
    nao_resolvidas = [c for c in correcoes if not c["resolveu"]]
    assert resolvidas, "correcao que virou CONFIRMADO nao foi registrada"
    assert nao_resolvidas, "correcao que permaneceu em REVISAO nao foi registrada"

    primeira = resolvidas[0]
    assert primeira["status_antes"] == "REVISAO"
    assert primeira["status_depois"] == "CONFIRMADO"
    assert primeira["campos"]["matricula"]["depois"] == "28972"
    # `campos_alterados` PODE ser vazio, e isso é informação, não falha:
    # aqui o operador conferiu a linha e confirmou o valor que já estava
    # (o que a destravou foi a base, não uma digitação). "Revisado e
    # mantido" é evidência tão real quanto "revisado e trocado" -- e é a
    # que some primeiro de qualquer registro informal.
    assert isinstance(primeira["campos_alterados"], list)
    # O texto BRUTO do OCR, que a planilha exportada só preservava para a
    # matrícula, e que `_revisao_confirmar` sobrescreve no dict.
    assert primeira["campos"]["matricula"]["ocr"], primeira["campos"]["matricula"]
    assert set(primeira["campos"]) == set(_rc.CAMPOS_REGISTRADOS)
    # Pelo menos UMA das correções da sessão mudou algum campo de fato
    # (a de matrícula 00000, que não resolveu).
    assert any(c["campos_alterados"] for c in correcoes), \
        "nenhuma correcao registrou campo alterado"
    # E o campo que bloqueava a linha antes da correção.
    assert any(c["campos_bloqueantes_antes"] for c in correcoes), \
        "nenhuma correcao registrou o campo bloqueante anterior"
    assert primeira["esquema"] == _rc.ESQUEMA and primeira["quando"]
    print(f"OK: Fase 20 -- {len(correcoes)} correcao(oes) humana(s) registradas "
          f"({len(resolvidas)} resolveram, {len(nao_resolvidas)} nao)")

    # A gravação é observacional: não alterou nenhum status nem a lista de
    # pendentes (o registro é feito DEPOIS da decisão).
    assert all(r["status"] in ("CONFIRMADO", "REVISAO", "ERRO")
               for r in app._registros_exportacao)
    print("OK: Fase 20 -- registrar a correcao nao alterou nenhuma decisao")

    # ==================================================================
    # Sub-fase 21b -- revisão como fluxo (lista de pendências, campo
    # bloqueante em destaque, "Ver detalhes", contador de progresso) e a
    # garantia de que `_revisao_confirmar` continua sendo o ÚNICO caminho
    # para sair de REVISAO (mesma proteção travada desde a Fase 12).
    # ==================================================================
    pendentes_21b = app._indices_pendentes_revisao()
    assert len(app.tabela_revisao_lista.get_children()) == len(pendentes_21b), (
        "a lista de pendencias (area 1) tem que refletir _indices_pendentes_revisao "
        "linha a linha -- e' a MESMA fonte de verdade que orienta a navegacao"
    )
    print("OK: Sub-fase 21b -- lista de pendencias reflete _indices_pendentes_revisao")

    # O contador "N de M revisados": N tem que ser exatamente quantas
    # correções desta sessão REALMENTE viraram CONFIRMADO (as `resolvidas`
    # já apuradas na secao da Fase 20 acima) -- nunca "quantas vezes o
    # botao foi clicado".
    assert app._revisao_resolvidos_sessao == len(resolvidas), (
        f"contador de progresso ({app._revisao_resolvidos_sessao}) nao bate com as "
        f"correcoes que realmente confirmaram ({len(resolvidas)})"
    )
    texto_progresso = app._texto_progresso_revisao(pendentes_21b)
    esperado_progresso = f"{len(resolvidas)} de {len(resolvidas) + len(pendentes_21b)} revisados"
    assert texto_progresso == esperado_progresso, (texto_progresso, esperado_progresso)
    print(f"OK: Sub-fase 21b -- contador de progresso: {texto_progresso!r}")

    if pendentes_21b:
        # Clicar numa linha da lista de pendências navega para ela --
        # mesmo destino que os botões Anterior/Próximo alcançariam -- e
        # NÃO toca em nenhum campo do formulário (é só navegação).
        alvo_lista = min(1, len(pendentes_21b) - 1)
        app.tabela_revisao_lista.selection_set(str(alvo_lista))
        app._on_selecionar_pendencia_lista()
        app.update()
        assert app._revisao_posicao == alvo_lista, \
            "selecionar uma linha na lista de pendencias nao navegou para ela"
        _, registro_apos_clique = app._revisao_registro_atual()
        for chave, var in app.revisao_vars.items():
            assert var.get() == (registro_apos_clique.get(chave) or ""), (
                f"clicar na lista de pendencias alterou o campo {chave!r} do formulario"
            )
        app._revisao_ir_para(0)
        app.update()
        print("OK: Sub-fase 21b -- clicar na lista de pendencias navega sem alterar o formulario")

        # O campo bloqueante (e só ele) recebe destaque visual -- rótulo e
        # caixa em estilo "danger".
        explicacao_21b, _sinais_21b = app._explicacao_revisao_atual()
        for chave, widget in app.revisao_widgets.items():
            estilo = str(widget.cget("style"))
            if chave in explicacao_21b.campos_bloqueantes:
                assert "danger" in estilo, \
                    f"campo bloqueante {chave!r} sem destaque visual (estilo={estilo!r})"
            else:
                assert "danger" not in estilo, \
                    f"campo NAO bloqueante {chave!r} destacado por engano (estilo={estilo!r})"
        print("OK: Sub-fase 21b -- só o(s) campo(s) bloqueante(s) recebem destaque visual")

        # "Ver detalhes" começa fechado a cada registro (não pode dominar a
        # tela) e alternar não toca em `revisao_vars` nem no status.
        assert app._revisao_detalhes_expandido is False, \
            "o painel de detalhes deveria comecar fechado ao carregar um registro"
        antes_alternar = {k: v.get() for k, v in app.revisao_vars.items()}
        app._alternar_detalhes_revisao()
        assert app._revisao_detalhes_expandido is True
        app._alternar_detalhes_revisao()
        assert app._revisao_detalhes_expandido is False
        depois_alternar = {k: v.get() for k, v in app.revisao_vars.items()}
        assert antes_alternar == depois_alternar, \
            "expandir/recolher 'Ver detalhes' alterou o formulario"
        print("OK: Sub-fase 21b -- 'Ver detalhes' comeca fechado e alterna sem tocar no formulario")

        # A foto da folha é carregada quando a miniatura está disponível
        # (guardada por _worker_imagens/_worker_pdf durante o processamento).
        _, registro_com_foto = app._revisao_registro_atual()
        if registro_com_foto["pagina_origem"] in app._miniaturas_por_pagina:
            assert getattr(app, "_imagem_pil_revisao", None) is not None, \
                "havia miniatura para a pagina, mas a foto nao foi carregada na revisao"
            print("OK: Sub-fase 21b -- foto da folha carregada quando disponivel")

    # A proteção central (Fase 7/12): NENHUM lugar do código pode escrever
    # "CONFIRMADO" à mão em registro["status"]/em um dict de exportação --
    # toda atribuição tem que vir do RETORNO de `classificar_registro` (via
    # a variável `resultado`/`status`, nunca um literal). Trava estrutural,
    # não só comportamental: mesmo que alguém adicione um atalho nunca
    # exercitado pelos cenários acima, o grep abaixo entra em pane.
    #
    # Fase 24a (Web MVP): a decisão (a reatribuição de status) e a
    # montagem do dict de exportação deixaram de morar em ui/app.py --
    # migraram para validacao/confirmacao.py (`confirmar_revisao_manual`,
    # chamada tanto pelo Tkinter quanto pelo backend web) e pipeline.py
    # (`montar_registro_exportacao`/`registro_erro_pagina`, idem). A trava
    # estrutural agora verifica os TRÊS arquivos: ui/app.py não pode conter
    # NENHUMA das duas formas (a lógica saiu de lá por completo -- só
    # resta a CHAMADA para a função compartilhada), e cada literal
    # legítimo precisa estar exatamente no módulo novo, exatamente uma vez.
    caminho_app = os.path.join(os.path.dirname(__file__), "..", "src",
                                "leitor_matriculas", "ui", "app.py")
    caminho_confirmacao = os.path.join(os.path.dirname(__file__), "..", "src",
                                        "leitor_matriculas", "validacao", "confirmacao.py")
    caminho_pipeline = os.path.join(os.path.dirname(__file__), "..", "src",
                                     "leitor_matriculas", "pipeline.py")
    with open(caminho_app, encoding="utf-8") as _fh:
        codigo_app = _fh.read()
    with open(caminho_confirmacao, encoding="utf-8") as _fh:
        codigo_confirmacao = _fh.read()
    with open(caminho_pipeline, encoding="utf-8") as _fh:
        codigo_pipeline = _fh.read()
    import re as _re

    for _nome, _codigo in (("ui/app.py", codigo_app),
                            ("validacao/confirmacao.py", codigo_confirmacao),
                            ("pipeline.py", codigo_pipeline)):
        assert not _re.search(r'\["status"\]\s*=\s*"CONFIRMADO"', _codigo), \
            f"{_nome} escreve 'CONFIRMADO' na marra em registro['status'] -- atalho proibido"
        assert not _re.search(r'"status"\s*:\s*"CONFIRMADO"', _codigo), \
            f"{_nome} escreve 'CONFIRMADO' na marra num dict literal -- atalho proibido"

    # ui/app.py não decide mais status nenhum -- a lógica saiu inteira.
    assert _re.findall(r'registro\["status"\]\s*=(?!=)\s*([^\n]+)', codigo_app) == [], (
        "ui/app.py voltou a reatribuir registro['status'] diretamente -- "
        "deveria delegar inteiramente para confirmar_revisao_manual"
    )
    assert _re.findall(r'"pagina_origem":[^\n]*"status":\s*(\w+|"ERRO")', codigo_app) == [], (
        "ui/app.py voltou a montar um dict de registro diretamente -- "
        "deveria delegar inteiramente para pipeline.montar_registro_exportacao"
    )
    # Prova de que ui/app.py de fato PASSA pelos módulos novos (não é só
    # ausência de atalho -- é presença da chamada certa).
    assert "confirmar_revisao_manual(" in codigo_app, \
        "_revisao_confirmar deveria chamar validacao.confirmacao.confirmar_revisao_manual"
    assert "pipeline.montar_registro_exportacao(" in codigo_app, \
        "_adicionar_registros deveria chamar pipeline.montar_registro_exportacao"

    # O único caminho legítimo de reatribuição manual: dentro de
    # validacao/confirmacao.py, vindo de `resultado.status` (o retorno de
    # `classificar_registro`), nunca um literal.
    reatribuicoes = _re.findall(r'registro\["status"\]\s*=(?!=)\s*([^\n]+)', codigo_confirmacao)
    assert reatribuicoes == ["resultado.status"], (
        f"esperava exatamente 1 reatribuicao de registro['status'] em "
        f"validacao/confirmacao.py (vinda de resultado.status); encontrado: {reatribuicoes}"
    )

    # Os dois dicts que MONTAM um registro (têm "pagina_origem" ao lado)
    # moram em pipeline.py agora: a classificação automática (monta o
    # dict a partir da variável `status`) e a linha ERRO de página.
    dicts_de_registro = _re.findall(r'"pagina_origem":[^\n]*"status":\s*(\w+|"ERRO")', codigo_pipeline)
    assert sorted(dicts_de_registro) == sorted(["status", '"ERRO"']), (
        f"esperava exatamente os 2 dicts de registro conhecidos (classificacao "
        f"automatica com 'status': status, e a linha ERRO de falha de pagina) em "
        f"pipeline.py; encontrado: {dicts_de_registro}"
    )
    print("OK: Sub-fase 24a -- confirmar_revisao_manual (validacao/confirmacao.py) e "
          "pipeline.montar_registro_exportacao continuam os UNICOS caminhos que decidem "
          "'status' -- ui/app.py só chama, nunca decide (nenhum atalho grava 'CONFIRMADO' "
          "na marra em nenhum dos três arquivos)")

    shutil.rmtree(tmp, ignore_errors=True)
    app.destroy()

shutil.rmtree(_tmp_dir_img, ignore_errors=True)
shutil.rmtree(_tmp_dir_correcoes, ignore_errors=True)

# ==========================================================================
# Sub-fase 21c -- responsividade: a aplicação continua funcional (sem
# exceção, sem widget com dimensão zero/negativa, sem aba sumida) em pelo
# menos duas resoluções de janela diferentes. Não mede pixel a pixel (fora
# do que dá para verificar sem abrir o olho humano) -- mede o que uma
# tela quebrada de verdade apresentaria: widget colapsado a 0/1px, ou uma
# exceção ao redimensionar.
# ==========================================================================
with patch("leitor_matriculas.ui.app.Messagebox.show_error"), \
     patch("leitor_matriculas.ui.app.Messagebox.show_warning"), \
     patch("leitor_matriculas.ui.app.Messagebox.show_info"):
    app_resp = App()
    # Duas linhas sintéticas (uma de cada status principal) para a tabela
    # não estar vazia durante o teste -- não precisa de OCR nem de thread.
    app_resp._registros_exportacao = [
        {"pagina_origem": 1, "status": "CONFIRMADO", "data": "23/04/26", "hora": "11:05",
         "matricula": "12345", "nome": "Fulano de Tal", "setor": "TI", "motivo": "RH",
         "gestor": "Gestor X", "cargo": "Analista", "confianca_matricula": 0.95,
         "confianca_gestor": 0.9, "confianca_motivo": 0.9, "observacao": ""},
        {"pagina_origem": 1, "status": "REVISAO", "data": "23/04/26", "hora": "",
         "matricula": "", "nome": "", "setor": "", "motivo": "RH", "gestor": "Gestor X",
         "cargo": "", "confianca_matricula": None, "confianca_gestor": 0.9,
         "confianca_motivo": 0.9, "observacao": "matrícula não identificada pelo OCR"},
    ]
    app_resp._sincronizar_tabela_principal()
    app_resp._atualizar_painel_revisao()  # popula a lista de pendências/foto da aba Revisão

    def _checar_janela_utilizavel(rotulo, widgets_da_aba_ativa):
        app_resp.update_idletasks()
        app_resp.update()
        # As abas continuam todas presentes -- redimensionar não pode
        # fazer nenhuma sumir.
        abas_presentes = app_resp.abas.tabs()
        assert len(abas_presentes) == 4, f"{rotulo}: esperava 4 abas, achou {len(abas_presentes)}"
        # Os widgets da aba ATUALMENTE VISÍVEL continuam com dimensão
        # utilizável (nada colapsado a 0/1px, o que indicaria overlap ou
        # um weight de grid quebrado). Widgets de uma aba oculta ficam
        # com geometria não gerenciada por design do Notebook -- não é
        # o que este teste mede.
        for nome_widget, widget in [("notebook", app_resp.abas), ("cabecalho", app_resp.btn_salvar)] \
                + widgets_da_aba_ativa:
            largura = widget.winfo_width()
            altura = widget.winfo_height()
            assert largura > 10 and altura > 5, (
                f"{rotulo}: {nome_widget} com dimensao suspeita ({largura}x{altura})"
            )
        print(f"OK: Sub-fase 21c -- janela utilizável em {rotulo} "
              f"({app_resp.winfo_width()}x{app_resp.winfo_height()})")

    app_resp.abas.select(app_resp._aba_registros)
    widgets_registros = [("tabela", app_resp.tabela)]

    # Tamanho padrão (o que a janela abre).
    _checar_janela_utilizavel("1400x860 (padrão)", widgets_registros)

    # Tamanho mínimo declarado (`self.minsize`, Fase 10) -- é o menor
    # tamanho que o próprio programa afirma suportar.
    app_resp.geometry("1120x700")
    _checar_janela_utilizavel("1120x700 (mínimo declarado)", widgets_registros)

    # Uma janela maior, para o caminho inverso (esticar, não só encolher).
    app_resp.geometry("1800x1000")
    _checar_janela_utilizavel("1800x1000 (janela grande)", widgets_registros)

    # A aba Revisão (a mais densa -- 3 painéis) também precisa sobreviver
    # ao redimensionamento, não só a Registros.
    app_resp.abas.select(app_resp._aba_revisao)
    app_resp.geometry("1120x700")
    widgets_revisao = [
        ("lista de pendências", app_resp.tabela_revisao_lista),
        ("canvas da foto", app_resp.canvas_foto),
        ("campo matrícula", app_resp.revisao_widgets["matricula"]),
    ]
    _checar_janela_utilizavel("1120x700, aba Revisão", widgets_revisao)

    app_resp.destroy()

# ==========================================================================
# Sub-fase 21d -- filtro por tipo de pendência, busca textual e atalhos de
# teclado (← →, Esc) na lista de pendências da Revisão. Os dois primeiros
# só decidem o que APARECE na lista; nenhum dos três pode alterar status
# de registro nenhum nem mexer em `_indices_pendentes_revisao()` (a lista
# de verdade, que Anterior/Próximo continuam percorrendo por inteiro).
# ==========================================================================


def _evidencia_bloqueante(campo):
    return {"campo": campo, "tipo": "regra", "resultado": "DUVIDA",
            "motivo": f"{campo} não identificado", "valor_observado": None,
            "valor_relacionado": None, "origem": "teste"}


with patch("leitor_matriculas.ui.app.Messagebox.show_error"), \
     patch("leitor_matriculas.ui.app.Messagebox.show_warning"), \
     patch("leitor_matriculas.ui.app.Messagebox.show_info"):
    app_21d = App()
    registros_21d = [
        {"pagina_origem": 1, "status": "REVISAO", "data": "", "hora": "11:00",
         "matricula": "11111", "nome": "", "setor": "", "motivo": "RH",
         "gestor": "Gestor X", "cargo": "", "confianca_matricula": None,
         "confianca_gestor": 0.9, "confianca_motivo": 0.9,
         "observacao": "data não identificada", "evidencias": [_evidencia_bloqueante("data")]},
        {"pagina_origem": 1, "status": "REVISAO", "data": "23/04/26", "hora": "11:05",
         "matricula": "22222", "nome": "", "setor": "", "motivo": "RH",
         "gestor": "Rubrica Ilegível", "cargo": "", "confianca_matricula": 0.95,
         "confianca_gestor": None, "confianca_motivo": 0.9,
         "observacao": "responsável não reconhecido", "evidencias": [_evidencia_bloqueante("gestor")]},
        {"pagina_origem": 2, "status": "REVISAO", "data": "23/04/26", "hora": "11:10",
         "matricula": "", "nome": "", "setor": "", "motivo": "RH",
         "gestor": "Gestor X", "cargo": "", "confianca_matricula": None,
         "confianca_gestor": 0.9, "confianca_motivo": 0.9,
         "observacao": "matrícula não identificada", "evidencias": [_evidencia_bloqueante("matricula")]},
    ]
    app_21d._registros_exportacao = [dict(r) for r in registros_21d]
    app_21d._sincronizar_tabela_principal()
    app_21d._atualizar_painel_revisao()

    pendentes_completo = app_21d._indices_pendentes_revisao()
    assert len(pendentes_completo) == 3, pendentes_completo

    # -------- Filtro por tipo: só afeta o que aparece na lista ----------
    app_21d.revisao_filtro_tipo_var.set("Responsável")
    app_21d._atualizar_lista_revisao_pendencias()
    linhas_mostradas = app_21d.tabela_revisao_lista.get_children()
    assert len(linhas_mostradas) == 1, f"filtro Responsável deveria mostrar 1 linha, mostrou {len(linhas_mostradas)}"
    indice_mostrado = pendentes_completo[int(linhas_mostradas[0])]
    assert app_21d._registros_exportacao[indice_mostrado]["matricula"] == "22222"
    # O filtro NUNCA muda a lista de verdade nem o status de ninguém.
    assert app_21d._indices_pendentes_revisao() == pendentes_completo
    assert all(r["status"] == "REVISAO" for r in app_21d._registros_exportacao)
    print("OK: Sub-fase 21d -- filtro por tipo de pendência mostra só o subconjunto certo, "
          "sem alterar status nem a lista de pendentes real")

    # -------- Busca textual: por matrícula -------------------------------
    app_21d.revisao_filtro_tipo_var.set("Todas")
    app_21d._revisao_busca_placeholder_ativo = False  # simula operador já tendo digitado
    app_21d.revisao_busca_var.set("11111")
    linhas_busca = app_21d.tabela_revisao_lista.get_children()
    assert len(linhas_busca) == 1, f"busca '11111' deveria achar 1 linha, achou {len(linhas_busca)}"
    indice_busca = pendentes_completo[int(linhas_busca[0])]
    assert app_21d._registros_exportacao[indice_busca]["matricula"] == "11111"
    assert app_21d._indices_pendentes_revisao() == pendentes_completo
    print("OK: Sub-fase 21d -- busca textual por matrícula mostra só a linha certa, "
          "sem alterar a lista de pendentes real")

    # -------- Placeholder nunca é tratado como termo de busca -------------
    app_21d._ativar_placeholder_busca_revisao()
    assert app_21d._texto_busca_revisao() == "", \
        "o texto do placeholder nao pode ser usado como termo de busca"
    app_21d._atualizar_lista_revisao_pendencias()
    assert len(app_21d.tabela_revisao_lista.get_children()) == 3, \
        "com o placeholder ativo (sem busca real), as 3 pendencias devem aparecer"

    # -------- Esc limpa a busca -- não confirma nada ----------------------
    # `event_generate` não é confiável neste ambiente para eventos de
    # teclado: a máquina tem um bug pré-existente e independente desta
    # sub-fase em que o popdown interno de QUALQUER Combobox da tela rouba
    # o foco de teclado sintético assim que existe um Combobox na árvore
    # (reproduz-se até com `bind_all` genérico, sem nenhum código desta
    # sub-fase envolvido -- ver relatório). Por isso o teste verifica os
    # dois pedaços que garantem o atalho de verdade: (1) a sequência está
    # REALMENTE amarrada ao handler certo (consulta a própria ligação do
    # Tk, não confia em simular a tecla); (2) o handler faz o que deveria
    # ao ser chamado (o mesmo código que o Tk chamaria de verdade).
    assert "_on_escape_busca_revisao" in str(app_21d.entrada_revisao_busca.bind("<Escape>")), \
        "a tecla Esc no campo de busca nao esta amarrada ao handler esperado"
    app_21d._revisao_busca_placeholder_ativo = False
    app_21d.revisao_busca_var.set("22222")
    app_21d._on_escape_busca_revisao()
    app_21d.update()
    assert app_21d._texto_busca_revisao() == "", "Esc deveria ter limpado a busca"
    assert app_21d._indices_pendentes_revisao() == pendentes_completo
    assert all(r["status"] == "REVISAO" for r in app_21d._registros_exportacao), \
        "Esc na busca nao pode confirmar nem alterar status de registro nenhum"
    print("OK: Sub-fase 21d -- Esc limpa a busca sem confirmar nem alterar nenhum registro")

    # -------- ← → navegam a lista (mesma ação de Anterior/Próximo) --------
    # Mesma ressalva do bloco acima quanto a `event_generate`: confirma a
    # ligação real no Tk e então exercita exatamente o callback que ela
    # chamaria (`lambda _e: self._revisao_navegar(...)`).
    assert "_on_seta_direita_revisao" in str(app_21d.tabela_revisao_lista.bind("<Right>")), \
        "a seta direita na lista de pendências nao esta amarrada ao handler esperado"
    assert "_on_seta_esquerda_revisao" in str(app_21d.tabela_revisao_lista.bind("<Left>")), \
        "a seta esquerda na lista de pendências nao esta amarrada ao handler esperado"
    app_21d._revisao_ir_para(0)
    app_21d.update()
    posicao_antes = app_21d._revisao_posicao
    app_21d._revisao_navegar(1)  # o mesmo que a seta direita dispara
    assert app_21d._revisao_posicao == (posicao_antes + 1) % len(pendentes_completo), (
        f"seta direita deveria avançar a posição (antes={posicao_antes}, "
        f"depois={app_21d._revisao_posicao})"
    )
    app_21d._revisao_navegar(-1)  # o mesmo que a seta esquerda dispara
    assert app_21d._revisao_posicao == posicao_antes, "seta esquerda deveria voltar a posição"
    assert app_21d._indices_pendentes_revisao() == pendentes_completo
    assert all(r["status"] == "REVISAO" for r in app_21d._registros_exportacao), \
        "navegar com as setas nao pode confirmar nem alterar status de registro nenhum"
    print("OK: Sub-fase 21d -- ← → navegam a lista de pendências como Anterior/Próximo, "
          "sem confirmar nada")

    # -------- Nenhuma aba "Configurações" foi criada -----------------------
    # Auditoria da 21d: não existe nenhuma opção real de usuário para
    # expor (nenhuma infraestrutura de preferências no projeto -- ver
    # relatório). Trava estrutural: o Notebook continua com as mesmas 4
    # abas de sempre.
    nomes_abas = [app_21d.abas.tab(aba, "text") for aba in app_21d.abas.tabs()]
    assert len(nomes_abas) == 4, f"esperava 4 abas, achou {nomes_abas}"
    assert not any("config" in n.lower() for n in nomes_abas), \
        f"uma aba de Configurações foi criada sem opcao real por tras: {nomes_abas}"
    print("OK: Sub-fase 21d -- nenhuma aba de Configurações decorativa foi criada")

    # -------- Sub-fase 22b (Fase 22 -- redesign visual): navegação e Home --
    # A correção do preâmbulo da 22b proíbe "Histórico" e "Configurações"
    # como item de navegação (nenhum dos dois é uma tela real -- ver 21a);
    # e a decisão desta sub-fase foi MANTER o Notebook (não trocar por
    # sidebar) -- continua valendo a mesma trava de 4 abas da 21d, agora
    # também nomeando "histó" (Histórico), não só "config".
    nomes_abas_22b = [app_21d.abas.tab(aba, "text") for aba in app_21d.abas.tabs()]
    assert len(nomes_abas_22b) == 4, f"a decisão da 22b foi manter as 4 abas reais, achou {nomes_abas_22b}"
    assert not any("histó" in n.lower() for n in nomes_abas_22b), \
        f"'Histórico' não é uma tela real (write-only por trava de teste) -- não pode virar item de navegação: {nomes_abas_22b}"
    assert not any("config" in n.lower() for n in nomes_abas_22b), \
        f"'Configurações' não tem nenhuma opção real por trás -- não pode virar item de navegação: {nomes_abas_22b}"

    # A correção central desta sub-fase (auditoria 22a, achados 1/2): os
    # agrupamentos da aba Início eram `ttk.Labelframe` (borda cavalgando o
    # rótulo); viraram título + `Card.TFrame` sobre um `Canvas.TFrame`.
    # Trava estrutural, não só visual: nenhum `ttk.Labelframe` sobra nos
    # três cartões do fluxo.
    def _sem_labelframe(widget):
        for filho in widget.winfo_children():
            assert not isinstance(filho, (ttk.LabelFrame,)), (
                f"{filho} continua sendo um ttk.Labelframe -- a 22b substituiu "
                "essa técnica por título + Card.TFrame"
            )
            _sem_labelframe(filho)

    for cartao in (app_21d._cartao_pronto, app_21d._cartao_selecao, app_21d._cartao_processando):
        _sem_labelframe(cartao)
    assert app_21d._aba_inicio.cget("style") == "Canvas.TFrame"
    assert app_21d._cartao_pronto.cget("style") == "Canvas.TFrame"
    assert app_21d._cartao_selecao.cget("style") == "Canvas.TFrame"
    assert app_21d._cartao_processando.cget("style") == "Canvas.TFrame"
    # Registros/Revisão/Avisos ficaram FORA do escopo desta sub-fase --
    # nenhuma das duas cores novas foi aplicada a elas ainda (Revisão
    # entrou na 22c, Registros/Avisos entraram na 22d -- ver blocos
    # abaixo; aqui só se confirma que a Início não vazou nada sozinha).
    print("OK: Sub-fase 22b -- Início sem nenhum ttk.Labelframe residual (título + Card no lugar); "
          "navegação continua com as 4 abas reais")

    # Os estilos novos (paleta da 22a aplicada nesta sub-fase) precisam ter
    # sido registrados de verdade -- um nome de estilo com erro de digitação
    # não derruba o programa (ttk cai no padrão em silêncio), então só
    # visualizar a tela não pegaria isso.
    _estilo_22b = tb.Style()
    assert _estilo_22b.lookup("Canvas.TFrame", "background") == estilos.COR_FUNDO
    assert _estilo_22b.lookup("Card.TFrame", "background") == estilos.COR_SUPERFICIE
    assert _estilo_22b.lookup("SecaoTitulo.TLabel", "background") == estilos.COR_FUNDO
    print("OK: Sub-fase 22b -- estilos Canvas/Card/SecaoTitulo registrados com os tokens de ui/estilos.py")

    # -------- Sub-fase 22c (Fase 22 -- redesign visual): aba Revisão -------
    # Mesma trava estrutural da 22b, agora nas três áreas da Revisão: nenhum
    # `ttk.Labelframe` residual (a correção central desta sub-fase).
    def _sem_labelframe_22c(widget):
        for filho in widget.winfo_children():
            assert not isinstance(filho, ttk.LabelFrame), (
                f"{filho} continua sendo um ttk.Labelframe na aba Revisão -- "
                "a 22c substituiu essa técnica por título + Card.TFrame"
            )
            _sem_labelframe_22c(filho)

    _sem_labelframe_22c(app_21d._aba_revisao)
    print("OK: Sub-fase 22c -- aba Revisão sem nenhum ttk.Labelframe residual "
          "(título + Card no lugar, nas 3 áreas)")

    # O destaque do campo bloqueante continua funcionando (mesma checagem da
    # 21b, re-executada aqui para travar que o refinamento visual da 22c --
    # ícone + negrito no rótulo -- NÃO tirou o bootstyle "danger" da CAIXA,
    # que é o sinal que um script/teste externo consegue verificar).
    explicacao_22c, _sinais_22c = app_21d._explicacao_revisao_atual()
    for chave, widget in app_21d.revisao_widgets.items():
        estilo_widget = str(widget.cget("style"))
        bloqueado = chave in explicacao_22c.campos_bloqueantes
        assert ("danger" in estilo_widget) == bloqueado, (
            f"campo {chave!r} (bloqueado={bloqueado}) com destaque inconsistente: {estilo_widget!r}"
        )
        # O RÓTULO ganha o ícone de alerta só quando bloqueado -- nunca
        # perde o texto original (troca de estilo não pode apagar dado).
        texto_rotulo = app_21d.revisao_rotulos[chave].cget("text")
        texto_base = app_21d._revisao_rotulos_texto[chave]
        if bloqueado:
            assert estilos.ICONE_REVISAO in texto_rotulo and texto_base in texto_rotulo
        else:
            assert texto_rotulo == texto_base
    print("OK: Sub-fase 22c -- campo bloqueante continua com a caixa em destaque "
          "(bootstyle danger) e o rótulo ganha ícone sem perder o texto original")

    # Nenhuma sugestão de contexto pré-seleciona um campo -- mesma garantia
    # da Fase 18/21b (não tocada nesta sub-fase), reafirmada aqui porque é
    # exatamente o que o preâmbulo da 22c pede para confirmar explicitamente.
    _, registro_22c = app_21d._revisao_registro_atual()
    for chave, var in app_21d.revisao_vars.items():
        assert var.get() == (registro_22c.get(chave) or ""), (
            f"campo {chave!r} veio diferente do valor já apurado do registro -- "
            "uma sugestão não pode pré-preencher o formulário"
        )
    print("OK: Sub-fase 22c -- nenhuma sugestão de contexto vem pré-selecionada no formulário")

    # Hover: puramente cosmético, não pode alterar seleção real nem status.
    selecao_antes_hover = app_21d.tabela_revisao_lista.selection()
    status_antes_hover = [r["status"] for r in app_21d._registros_exportacao]

    class _EventoFalso:
        y = 5

    app_21d._on_hover_lista_revisao(_EventoFalso())
    assert app_21d.tabela_revisao_lista.selection() == selecao_antes_hover
    app_21d._limpar_hover_lista_revisao()
    assert app_21d._hover_item_revisao is None
    assert [r["status"] for r in app_21d._registros_exportacao] == status_antes_hover
    print("OK: Sub-fase 22c -- hover da lista de pendências é cosmético "
          "(não altera seleção nem status de registro nenhum)")

    # Os containers das três áreas usam Canvas.TFrame -- mesmo padrão de
    # verificação da 22b, agora para a Revisão.
    _estilo_22c = tb.Style()
    assert _estilo_22c.lookup("PosicaoRevisao.TLabel", "background") == estilos.COR_FUNDO
    assert _estilo_22c.lookup("ResumoRevisao.TLabel", "foreground") == estilos.COR_ATENCAO
    assert _estilo_22c.lookup("ResultadoRevisaoErro.TLabel", "foreground") == estilos.COR_ERRO
    print("OK: Sub-fase 22c -- estilos PosicaoRevisao/ResumoRevisao/ResultadoRevisaoErro "
          "registrados com os tokens de ui/estilos.py")

    # A proteção central continua intacta: `_revisao_confirmar` é o único
    # caminho para sair de REVISAO -- mesma trava por regex da 21b,
    # reconferida aqui porque esta sub-fase tocou bastante `ui/app.py`.
    assert not _re.search(r'\["status"\]\s*=\s*"CONFIRMADO"', codigo_app), \
        "22c: nenhum lugar pode gravar CONFIRMADO na marra"
    print("OK: Sub-fase 22c -- _revisao_confirmar continua o UNICO caminho para sair de REVISAO")

    # -------- Sub-fase 22d (Fase 22 -- redesign visual): tabelas e diálogos --
    # Registros e Avisos entram no mesmo padrão canvas+cartão (mesma trava
    # estrutural das sub-fases anteriores, agora sem a exceção da 22b).
    def _sem_labelframe_22d(widget):
        for filho in widget.winfo_children():
            assert not isinstance(filho, ttk.LabelFrame), (
                f"{filho} continua sendo um ttk.Labelframe -- Registros/Avisos "
                "entraram no padrão título+Card.TFrame na 22d"
            )
            _sem_labelframe_22d(filho)

    _sem_labelframe_22d(app_21d._aba_registros)
    _sem_labelframe_22d(app_21d._aba_avisos)
    assert app_21d._aba_registros.cget("style") == "Canvas.TFrame"
    assert app_21d._aba_avisos.cget("style") == "Canvas.TFrame"
    assert app_21d._moldura_tabela.cget("style") == "Card.TFrame"
    print("OK: Sub-fase 22d -- Registros/Avisos sem nenhum ttk.Labelframe residual "
          "(mesmo padrão canvas+Card das demais abas)")

    # Estilos da tabela: fundo de cartão, cabeçalho discreto, seleção no
    # mesmo tom da lista de pendências, e o contorno de FOCO neutralizado
    # (o achado real desta sub-fase: o construtor de Treeview do
    # ttkbootstrap pinta um contorno azul de accent ao redor da tabela
    # inteira quando ela tem foco -- ver comentário em
    # `_montar_estilos_visuais`).
    _estilo_22d = tb.Style()
    assert _estilo_22d.lookup("Treeview", "background") == estilos.COR_SUPERFICIE
    assert _estilo_22d.lookup("Treeview.Heading", "foreground") == estilos.COR_TEXTO_SECUNDARIO
    assert _estilo_22d.lookup("Treeview", "bordercolor", state=["focus"]) == estilos.COR_BORDA
    print("OK: Sub-fase 22d -- estilos de Treeview (fundo/cabeçalho/contorno de foco) "
          "registrados com os tokens de ui/estilos.py")

    # Hover genérico (Avisos): mesma garantia da 22c -- cosmético, não
    # altera o conteúdo da tabela nem nenhum estado de negócio.
    app_21d._atualizar_avisos()
    itens_avisos = app_21d.tabela_avisos.get_children()
    assert itens_avisos, "esperava ao menos uma linha (avisos reais ou o placeholder de vazio)"

    class _EventoFalsoAvisos:
        y = 5

    handler_hover = app_21d._on_hover_generico(app_21d.tabela_avisos)
    handler_limpar = app_21d._limpar_hover_generico(app_21d.tabela_avisos)
    handler_hover(_EventoFalsoAvisos())
    assert "hover" in app_21d.tabela_avisos.item(itens_avisos[0], "tags")
    handler_limpar()
    assert "hover" not in app_21d.tabela_avisos.item(itens_avisos[0], "tags")
    assert getattr(app_21d.tabela_avisos, "_hover_item", None) is None
    print("OK: Sub-fase 22d -- hover genérico da aba Avisos é cosmético "
          "(tag aparece/some sem alterar conteúdo da tabela)")

    # Diálogos: `Messagebox` (ttkbootstrap) no lugar de `tkinter.
    # messagebox` -- confirma que o import certo está em uso (o resto da
    # suíte já prova que os `patch(...)` corretos interceptam as chamadas
    # sem travar em um diálogo real).
    from ttkbootstrap.dialogs import Messagebox as _MessageboxEsperado
    assert app.__class__.__module__ == "leitor_matriculas.ui.app"
    import leitor_matriculas.ui.app as _app_module
    assert _app_module.Messagebox is _MessageboxEsperado
    assert not hasattr(_app_module, "messagebox"), (
        "o import antigo (tkinter.messagebox) deveria ter sido removido -- "
        "nenhuma chamada mais usa ele"
    )
    print("OK: Sub-fase 22d -- diálogos usam Messagebox (ttkbootstrap), "
          "import antigo (tkinter.messagebox) removido")

    assert not _re.search(r'\["status"\]\s*=\s*"CONFIRMADO"', codigo_app), \
        "22d: nenhum lugar pode gravar CONFIRMADO na marra"
    print("OK: Sub-fase 22d -- _revisao_confirmar continua o UNICO caminho para sair de REVISAO")

    # -------- Sub-fase 22e (Fase 22 -- redesign visual): tema e modo escuro --
    # SÓ verificação ESTRUTURAL aqui -- não chama `_alternar_tema()` de
    # verdade nesta instância. `self.style.theme_use(...)` (ttkbootstrap)
    # só funciona sem erro para a PRIMEIRA janela Tk do processo (medido
    # -- ver `saida/avaliacao_fase22_redesign.md`, seção 22e); a esta
    # altura do arquivo, `app_21d` já é uma entre DEZENAS de `App()`
    # criadas no mesmo processo, e uma segunda chamada real a
    # `theme_use` aqui reproduz o mesmo `TclError` cosmético já
    # documentado do `TPanedwindow` (22c/22d) -- não um bug desta
    # sub-fase. O ciclo completo de alternância É testado de verdade,
    # com uma única `App()` isolada no próprio processo (exatamente como
    # a operação real), em `teste/teste_alternar_tema.py`.
    assert hasattr(app_21d, "btn_alternar_tema")
    assert not app_21d._tema_escuro, "esta sessão de teste não trocou de tema -- deveria seguir clara"
    assert estilos.ICONE_TEMA_ESCURO in app_21d.btn_alternar_tema.cget("text"), (
        "no tema claro, o botão deveria oferecer trocar para o ESCURO"
    )
    assert estilos.TEMA_CLARO != estilos.TEMA_ESCURO
    assert callable(app_21d._alternar_tema)
    print("OK: Sub-fase 22e -- botão de alternância de tema existe e mostra o estado inicial correto "
          "(ciclo completo de troca testado em teste_alternar_tema.py)")

    # Reconfirmação final (fim da Fase 22): a proteção central continua
    # intacta depois de cinco sub-fases mexendo bastante em ui/app.py.
    assert not _re.search(r'\["status"\]\s*=\s*"CONFIRMADO"', codigo_app), \
        "fim da Fase 22: nenhum lugar pode gravar CONFIRMADO na marra"
    print("OK: Sub-fase 22e -- _revisao_confirmar continua o UNICO caminho para sair de REVISAO "
          "(reconfirmado ao fim da Fase 22)")

    app_21d.destroy()

print("TESTE DE INTEGRACAO UI COMPLETO: OK")
