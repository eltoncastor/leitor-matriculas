"""
teste/teste_data_manager.py

Teste do DataManager, em duas partes:

PARTE A — bases reais do projeto (dados/Colaboradores.xlsx, Motivos.xlsx,
Gestores.xlsx). Se esses arquivos ainda não tiverem sido colocados na pasta
`dados/`, este teste NÃO inventa dados: ele apenas informa isso claramente
e segue para a parte B.

PARTE B — fixtures SINTÉTICAS (criadas e destruídas em uma pasta temporária,
só durante o teste). Servem para provar que a lógica de leitura do
DataManager está correta — preservação de zero à esquerda, coluna de
matrícula ausente, XLSX vazio, pasta ausente — independente de já termos ou
não os arquivos reais.

Uso:
    python teste\\teste_data_manager.py
    python teste\\teste_data_manager.py 30244      (testa uma matrícula real específica)
"""

import os
import sys
import shutil
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from leitor_matriculas.dados.data_manager import DataManager  # noqa: E402


def parte_a_bases_reais():
    print("=" * 70)
    print("PARTE A — bases reais do projeto (pasta dados/)")
    print("=" * 70)

    dm = DataManager()

    if dm.avisos:
        print("Avisos ao carregar as bases reais:")
        for aviso in dm.avisos:
            print(f"  - {aviso}")
    else:
        print("Nenhum aviso — as três bases foram carregadas sem problemas.")

    print()
    print("Resumo:", dm.resumo_status())

    if not dm.colaboradores_disponivel:
        print()
        print(
            "Colaboradores.xlsx real ainda não está disponível (ou não foi "
            "possível lê-lo) — nada será inventado. Assim que o arquivo "
            "real for colocado em dados/Colaboradores.xlsx, rode este teste "
            "de novo."
        )
        return

    matricula_teste = sys.argv[1] if len(sys.argv) > 1 else None
    if matricula_teste:
        print()
        print(f"Consultando matrícula informada: {matricula_teste!r}")
        resultado = dm.buscar_colaborador(matricula_teste)
        print(resultado if resultado else "Não encontrada na base - revisar")
    else:
        print()
        print(
            "Dica: rode 'python teste\\teste_data_manager.py <matricula>' "
            "para testar uma consulta real, por exemplo:"
        )
        print("  python teste\\teste_data_manager.py 30244")

    if dm.motivos_disponivel:
        print()
        print(f"Motivos carregados ({len(dm.motivos)}): {dm.listar_motivos()}")
    if dm.gestores_disponivel:
        print()
        print(f"Gestores carregados ({len(dm.gestores)}): {dm.listar_gestores()}")


def parte_b_fixtures_sinteticas():
    print()
    print("=" * 70)
    print("PARTE B — fixtures sintéticas (não são dados reais, só testam a lógica)")
    print("=" * 70)

    import openpyxl

    tmp = tempfile.mkdtemp()
    try:
        pasta_dados = os.path.join(tmp, "dados")
        os.makedirs(pasta_dados)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Matrícula", "Nome", "Cargo", "Setor"])
        ws.append(["00312", "Fulano de Teste", "Analista", "TI"])
        ws.append([30244, "Beltrano de Teste", "Auxiliar", "Prevenção"])
        wb.save(os.path.join(pasta_dados, "Colaboradores.xlsx"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Motivo"])
        ws.append(["Horário Negado"])
        ws.append(["RH"])
        wb.save(os.path.join(pasta_dados, "Motivos.xlsx"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Gestor Teste 1"])
        ws.append(["Gestor Teste 2"])
        wb.save(os.path.join(pasta_dados, "Gestores.xlsx"))

        dm = DataManager(pasta_dados=pasta_dados)

        print("Teste 1 (carregamento):", "OK" if not dm.avisos else f"FALHOU -> {dm.avisos}")
        assert not dm.avisos

        print("Teste 2 (buscar_colaborador existente):", end=" ")
        colaborador = dm.buscar_colaborador("00312")
        assert colaborador and colaborador["nome"] == "Fulano de Teste"
        print("OK ->", colaborador)

        print("Teste 3 (zero à esquerda preservado):", end=" ")
        assert dm.buscar_colaborador("00312")["matricula"] == "00312"
        print("OK")

        print("Teste 4 (matrícula inexistente):", end=" ")
        assert dm.buscar_colaborador("99999") is None
        print("OK -> None")

        print("Teste 5 (motivos):", end=" ")
        assert dm.listar_motivos() == ["Horário Negado", "RH"]
        print("OK ->", dm.listar_motivos())

        print("Teste 6 (gestores):", end=" ")
        assert dm.listar_gestores() == ["Gestor Teste 1", "Gestor Teste 2"]
        print("OK ->", dm.listar_gestores())

        # --- casos de borda ---
        pasta_vazia = os.path.join(tmp, "dados_vazio")
        os.makedirs(pasta_vazia)
        wb_vazio = openpyxl.Workbook()
        wb_vazio.save(os.path.join(pasta_vazia, "Colaboradores.xlsx"))
        wb_vazio.save(os.path.join(pasta_vazia, "Motivos.xlsx"))
        wb_vazio.save(os.path.join(pasta_vazia, "Gestores.xlsx"))
        dm_vazio = DataManager(pasta_dados=pasta_vazia)
        print("Teste extra (XLSX vazio não derruba o programa):", end=" ")
        assert len(dm_vazio.avisos) == 3 and not dm_vazio.colaboradores_disponivel
        print("OK")

        pasta_sem_coluna = os.path.join(tmp, "dados_sem_coluna")
        os.makedirs(pasta_sem_coluna)
        wb_sc = openpyxl.Workbook()
        ws_sc = wb_sc.active
        ws_sc.append(["Nome", "Cargo"])
        wb_sc.save(os.path.join(pasta_sem_coluna, "Colaboradores.xlsx"))
        dm_sc = DataManager(pasta_dados=pasta_sem_coluna)
        print("Teste extra (coluna de matrícula ausente vira aviso, não crash):", end=" ")
        assert any("Coluna de matrícula" in a for a in dm_sc.avisos)
        print("OK")

        dm_ausente = DataManager(pasta_dados=os.path.join(tmp, "nao_existe"))
        print("Teste extra (pasta dados/ ausente):", end=" ")
        assert len(dm_ausente.avisos) == 3
        print("OK")

        print()
        print("TODOS OS TESTES SINTÉTICOS PASSARAM.")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    parte_a_bases_reais()
    parte_b_fixtures_sinteticas()
